"""
dashboard_ruedas_MLL2.py
============================================================
Dashboard unificado — Control de Desgaste de Ruedas MLL2
============================================================
Une en una sola app de Streamlit:

  1. Alertas (según tendencia de desgaste proyectada)
  2. Evolución por eje + pronóstico de desgaste
  3. Heatmap de desgaste por tren (Izquierda / Derecha) — en vez del
     panel "Desgaste por rueda"
  4. Ranking — ruedas más próximas al límite
  5. Calendario de reperfilados

Incluye:
  - Selector de rango de fechas (calendario)
  - Carga de un Excel/CSV propio desde la misma app, sin depender
    de que el archivo esté siempre en la carpeta "datos"

Estructura de carpetas sugerida (ver README al final del archivo o
el mensaje de la conversación):

    Panel_Ruedas_MLL2/
      codigo/
        dashboard_ruedas_MLL2.py   <- este archivo
        requirements.txt
      datos/
        RUEDAS_FRENOS_PANTOGRAFO_04_2025_formateado.xlsx  (opcional)
      INICIAR.bat

Ejecutar (desde la carpeta "codigo"):
    streamlit run dashboard_ruedas_MLL2.py
"""

import glob
import os
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st

# ============================================================
# 1. CONFIGURACIÓN GENERAL
# ============================================================
LIMITES = {"Sd": 25.0, "Sh": 27.5, "QR": 6.5}
LIMITES_DESC = {
    "Sd": "Espesor de pestaña",
    "Sh": "Altura de pestaña",
    "QR": "Coeficiente de conicidad",
}
MARGEN_ALERTA_PCT = 0.10
NOMBRE_ARCHIVO_ESPERADO = "RUEDAS_FRENOS_PANTOGRAFO_04_2025_formateado.xlsx"

# Fechas de reperfilado — "todos" = todos los coches del tren,
# lista = reperfilado parcial (solo esos coches).
REPERFILADOS_DEFAULT = [
    {"vehiculo": 27, "fecha": "2024-09-17", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 14, "fecha": "2024-10-01", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 29, "fecha": "2025-02-05", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 28, "fecha": "2025-02-24", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 28, "fecha": "2026-02-11", "duracion_dias": 7, "alcance": ["M1"]},
    {"vehiculo": 26, "fecha": "2026-06-01", "duracion_dias": 7, "alcance": ["M1", "M2"]},
    {"vehiculo": 27, "fecha": "2026-05-26", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 14, "fecha": "2026-06-01", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 29, "fecha": "2026-06-22", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 26, "fecha": "2026-06-15", "duracion_dias": 7, "alcance": "todos"},
    {"vehiculo": 28, "fecha": "2026-06-30", "duracion_dias": 7, "alcance": "todos"},
]
# Lista "viva" que usan todas las funciones de este archivo. Se
# reemplaza en tiempo de ejecución con lo que el usuario edite en el
# panel "🗓️ Editar fechas de reperfilado" de la barra lateral.
REPERFILADOS = list(REPERFILADOS_DEFAULT)
# vehiculo|eje -> coche ("M1"/"M2"/...) — completar si se conoce la
# composición para afinar reperfilados parciales.
EJE_COCHE_MAP = {}

# Paleta visual (tema oscuro tipo "panel de control ferroviario")
PAL = {
    "bg": "#0f1216", "panel": "#161b20", "panel2": "#1c232b",
    "line": "#2b333c", "text": "#e9e7e1", "text_dim": "#98a1aa",
    "steel": "#4fb3e8", "amber": "#f3a938", "red": "#ef5468",
    "green": "#4fc98a", "violet": "#b18af0", "cyan": "#43d1c4",
}

COLORSCALE_HEAT = [
    [0.00, "#b71c1c"], [0.04, "#dc3545"], [0.05, "#e53935"],
    [0.06, "#ff6f00"], [0.08, "#ff8f00"], [0.10, "#ffa000"],
    [0.12, "#ffb300"], [0.14, "#ffc107"], [0.15, "#ffca28"],
    [0.16, "#cddc39"], [0.20, "#aed581"], [0.25, "#81c784"],
    [0.30, "#66bb6a"], [0.35, "#4caf50"], [0.40, "#43a047"],
    [0.45, "#2e7d32"], [0.50, "#00897b"], [0.55, "#009688"],
    [0.60, "#00acc1"], [0.65, "#00bcd4"], [0.70, "#29b6f6"],
    [0.75, "#03a9f4"], [0.80, "#1e88e5"], [0.85, "#1976d2"],
    [0.90, "#1565c0"], [0.95, "#0d47a1"], [1.00, "#0a237e"],
]


def layout_oscuro(fig, height=440, legend=True, hovermode="x unified"):
    """Aplica el tema oscuro estándar a una figura de Plotly."""
    fig.update_layout(
        paper_bgcolor=PAL["panel"], plot_bgcolor=PAL["panel"],
        font=dict(color=PAL["text"], family="IBM Plex Mono, monospace", size=11.5),
        height=height,
        hovermode=hovermode,
        legend=dict(orientation="h", y=1.1, bgcolor="rgba(0,0,0,0)") if legend else None,
        margin=dict(t=50, r=20, l=55, b=45),
    )
    fig.update_xaxes(gridcolor=PAL["line"], linecolor=PAL["line"], zeroline=False)
    fig.update_yaxes(gridcolor=PAL["line"], linecolor=PAL["line"], zeroline=False)
    return fig


# ============================================================
# 2. CARGA DE DATOS (Excel de la carpeta "datos" O archivo subido)
# ============================================================
def localizar_excel_por_defecto():
    """Busca el Excel/CSV en ../datos sin depender de la letra de unidad."""
    carpeta_script = Path(__file__).resolve().parent
    carpeta_datos = carpeta_script.parent / "datos"

    candidato = carpeta_datos / NOMBRE_ARCHIVO_ESPERADO
    if candidato.is_file():
        return candidato

    if carpeta_datos.is_dir():
        for patron in ("*.xlsx", "*.csv"):
            encontrados = [
                Path(f) for f in glob.glob(str(carpeta_datos / patron))
                if not Path(f).name.startswith("~$")
            ]
            if encontrados:
                return encontrados[0]

    # búsqueda amplia dentro de la carpeta del proyecto
    carpeta_proyecto = carpeta_script.parent
    for f in glob.glob(str(carpeta_proyecto / "**" / "*.xlsx"), recursive=True):
        if not Path(f).name.startswith("~$"):
            return Path(f)
    return None


def parsear_fecha(valor):
    if valor is None:
        return pd.NaT
    if isinstance(valor, (pd.Timestamp,)):
        return valor
    if hasattr(valor, "year") and hasattr(valor, "month"):
        return pd.Timestamp(valor)
    if isinstance(valor, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return pd.to_datetime(valor, format=fmt)
            except ValueError:
                continue
        return pd.to_datetime(valor, errors="coerce", dayfirst=True)
    return pd.NaT


def _parece_fila_de_datos(fila):
    """True si una fila 'parece' una medición real (vehiculo y eje
    numéricos + alguna fecha), sin importar en qué línea del archivo
    esté. Esto es lo que hace que la carga NO dependa de que los datos
    empiecen siempre en la misma fila (p. ej. si se agrega o quita una
    fila de encabezado en una actualización futura del archivo)."""
    if fila is None or len(fila) < 12:
        return False
    try:
        int(fila[1])
        int(fila[4])
    except (TypeError, ValueError):
        return False
    return fila[2] is not None


def _detectar_inicio_datos(filas):
    """Devuelve el índice de la primera fila que parece un dato real,
    explorando hasta las primeras 40 filas del archivo."""
    for i, fila in enumerate(filas[:40]):
        if _parece_fila_de_datos(fila):
            return i
    return None


def extraer_filas_crudas(fuente, es_csv):
    """Devuelve (filas_de_datos, n_filas_encabezado). Detecta
    automáticamente dónde empiezan los datos reales (en vez de asumir
    una fila fija), para que el archivo siga funcionando aunque cambie
    la cantidad de filas de encabezado en una actualización futura.

    Estructura de columnas esperada (igual en Excel y CSV):
    [accion, vehiculo, fecha, kms, eje,
     Sd_izq, QR_izq, Sh_izq, Rd_izq, Sd_der, QR_der, Sh_der, Rd_der]
    """
    if es_csv:
        df_raw = pd.read_csv(fuente, header=None, dtype=object)
        todas = df_raw.values.tolist()
    else:
        wb = openpyxl.load_workbook(fuente, data_only=True)
        hoja = "Pagina_1" if "Pagina_1" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[hoja]
        todas = list(ws.iter_rows(values_only=True))

    inicio = _detectar_inicio_datos(todas)
    if inicio is None:
        raise ValueError(
            "No se encontraron filas de datos reconocibles (se esperaba una "
            "columna de vehículo y una de eje numéricas). Verifica que el "
            "archivo mantenga la misma estructura de columnas que el "
            "original: [acción, vehículo, fecha, kms, eje, Sd_i, Qr_i, Sh_i, "
            "Rd_i, Sd_d, Qr_d, Sh_d, Rd_d]."
        )
    return todas[inicio:]


def _a_numero(valor):
    """Convierte a float de forma tolerante: strings vacíos, espacios
    o texto no numérico se vuelven NaN en vez de romper la carga."""
    if valor is None:
        return np.nan
    if isinstance(valor, str):
        valor = valor.strip().replace(",", ".")
        if valor == "":
            return np.nan
    try:
        return float(valor)
    except (TypeError, ValueError):
        return np.nan


def construir_dataframe(filas):
    """Convierte filas crudas en un DataFrame ancho, una fila por
    medición (vehiculo + fecha + eje). Cada fila se procesa de forma
    aislada: una fila mal formada se omite (se cuenta) en vez de
    detener la carga completa de todo el archivo."""
    registros = []
    n_omitidas = 0
    for fila in filas:
        try:
            if fila is None or len(fila) < 12:
                n_omitidas += 1
                continue
            vehiculo, fecha, kms, eje = fila[1], fila[2], fila[3], fila[4]
            if vehiculo is None or eje is None:
                n_omitidas += 1
                continue
            vehiculo = int(vehiculo)
            eje = int(eje)

            fecha_dt = parsear_fecha(fecha)
            if pd.isna(fecha_dt):
                n_omitidas += 1
                continue

            registros.append({
                "vehiculo": vehiculo, "fecha": fecha_dt,
                "kms": _a_numero(kms), "eje": eje,
                "Sd_izq": _a_numero(fila[5]), "QR_izq": _a_numero(fila[6]),
                "Sh_izq": _a_numero(fila[7]), "Rd_izq": _a_numero(fila[8]),
                "Sd_der": _a_numero(fila[9]), "QR_der": _a_numero(fila[10]),
                "Sh_der": _a_numero(fila[11]),
                "Rd_der": _a_numero(fila[12]) if len(fila) > 12 else np.nan,
            })
        except Exception:
            # cualquier fila con un problema inesperado se omite, nunca
            # rompe la carga de todo el archivo
            n_omitidas += 1
            continue

    if not registros:
        raise ValueError(
            "No se pudo extraer ninguna fila válida. Verifica que el archivo "
            "tenga la misma estructura de columnas que el Excel original."
        )

    df = pd.DataFrame(registros)

    # ---- Detección de duplicados (mismo vehículo + eje + fecha) ----
    # No se eliminan a ciegas: si el Excel trae dos filas con la misma
    # clave (p. ej. una corrección posterior), se conserva la ÚLTIMA
    # ocurrencia del archivo (se asume la más reciente/corregida) y se
    # informa la cantidad al usuario, para que pueda revisarlo si no
    # era lo esperado.
    n_duplicadas = int(df.duplicated(subset=["vehiculo", "eje", "fecha"], keep=False).sum())
    if n_duplicadas:
        df = df.drop_duplicates(subset=["vehiculo", "eje", "fecha"], keep="last")

    df = df.sort_values(["vehiculo", "fecha", "eje"]).reset_index(drop=True)
    return df, n_omitidas, n_duplicadas


@st.cache_data(show_spinner=False)
def cargar_desde_ruta(ruta_str, es_csv):
    filas = extraer_filas_crudas(ruta_str, es_csv)
    return construir_dataframe(filas)


@st.cache_data(show_spinner=False)
def cargar_desde_bytes(nombre, contenido, es_csv):
    import io
    buf = io.BytesIO(contenido)
    filas = extraer_filas_crudas(buf, es_csv)
    return construir_dataframe(filas)


def a_formato_largo(df_ancho):
    """Formato largo (vehiculo, fecha, eje, lado, variable, valor) para el heatmap."""
    registros = []
    for _, r in df_ancho.iterrows():
        for lado, suf in [("Izquierda", "izq"), ("Derecha", "der")]:
            for var in ["Sd", "QR", "Sh"]:
                val = r[f"{var}_{suf}"]
                if pd.isna(val):
                    continue
                registros.append({
                    "vehiculo": r["vehiculo"], "fecha": r["fecha"], "eje": r["eje"],
                    "lado": lado, "variable": var, "valor": float(val),
                })
    return pd.DataFrame(registros)


# ============================================================
# 2.1 DESGASTE DE DIÁMETRO (Rd) CADA 10,000 KM
# ============================================================
def calcular_desgaste_diametro(df, veh, eje):
    """Para un vehículo/eje: diferencia de diámetro (Rd) entre
    mediciones consecutivas, normalizada cada 10,000 km, separando
    rueda izquierda y derecha."""
    filas = []
    sub = df[(df["vehiculo"] == veh) & (df["eje"] == eje)]
    for lado, col in [("Izquierda", "Rd_izq"), ("Derecha", "Rd_der")]:
        serie = sub[["fecha", "kms", col]].dropna().sort_values("kms")
        serie = serie.drop_duplicates(subset="kms")
        if len(serie) < 2:
            continue
        prev = None
        for _, r in serie.iterrows():
            if prev is not None:
                delta_km = r["kms"] - prev["kms"]
                if delta_km <= 0:
                    prev = r
                    continue
                delta_rd = r[col] - prev[col]
                filas.append({
                    "lado": lado,
                    "fecha_inicio": prev["fecha"], "fecha_fin": r["fecha"],
                    "km_inicio": prev["kms"], "km_fin": r["kms"], "delta_km": delta_km,
                    "diametro_inicio_mm": prev[col], "diametro_fin_mm": r[col],
                    "delta_diametro_mm": delta_rd,
                    "desgaste_cada_10000km_mm": -(delta_rd / delta_km) * 10000,
                })
            prev = r
    return pd.DataFrame(filas)


def calcular_desgaste_diametro_global(df, vehiculos, ejes):
    """Igual que calcular_desgaste_diametro pero para todos los
    vehículos/ejes, usado para el ranking de mayor desgaste de diámetro."""
    partes = []
    for v in vehiculos:
        for e in ejes:
            d = calcular_desgaste_diametro(df, v, e)
            if not d.empty:
                d.insert(0, "eje", e)
                d.insert(0, "vehiculo", v)
                partes.append(d)
    if not partes:
        return pd.DataFrame()
    return pd.concat(partes, ignore_index=True)


def calcular_diferencia_izq_der(df, veh, eje):
    """Diferencia de diámetro entre rueda izquierda y derecha del
    mismo eje, en cada fecha de medición (indicador de desbalance)."""
    sub = df[(df["vehiculo"] == veh) & (df["eje"] == eje)]
    sub = sub[["fecha", "kms", "Rd_izq", "Rd_der"]].dropna().sort_values("fecha")
    if sub.empty:
        return sub
    sub = sub.copy()
    sub["diferencia_izq_der_mm"] = sub["Rd_izq"] - sub["Rd_der"]
    return sub


# ============================================================
# 3. LÓGICA DE REPERFILADOS / TENDENCIA / PRONÓSTICO
# ============================================================
def reperfilado_afecta_eje(evento, veh, eje):
    """IMPORTANTE: si el reperfilado es parcial (alcance = lista de
    coches) y no existe un mapeo eje→coche configurado para ese eje,
    NO se asume afectado. Asumir "sí" por defecto propagaría el evento
    a coches que en realidad no fueron reperfilados (viola el alcance
    real). Es preferible mostrar de más información sin filtrar, que
    filtrar datos reales de un coche que no tuvo la intervención."""
    if evento["vehiculo"] != veh:
        return False
    if evento["alcance"] == "todos":
        return True
    coche = EJE_COCHE_MAP.get(f"{veh}|{eje}")
    if coche is None:
        return False  # sin mapeo -> NO se asume afectado (seguro por defecto)
    return coche in evento["alcance"]


def hay_reperfilados_parciales_sin_mapear(vehiculos, ejes):
    """Detecta si existen reperfilados parciales para trenes cuyos ejes
    no tienen un mapeo a coche definido, para poder avisarle al
    usuario que configure el mapeo (de lo contrario esos eventos no se
    aplican a ningún eje, por el criterio seguro de arriba)."""
    veh_parciales = {r["vehiculo"] for r in REPERFILADOS if r["alcance"] != "todos"}
    faltantes = []
    for v in veh_parciales:
        if v not in vehiculos:
            continue
        ejes_v = [e for e in ejes if f"{v}|{e}" in EJE_COCHE_MAP]
        if not ejes_v:
            faltantes.append(v)
    return faltantes


def reperfilados_de_eje(veh, eje):
    """Devuelve los eventos de reperfilado que afectan a este eje, como
    tuplas (fecha_inicio, fecha_fin_estimada), usando la duración
    aproximada de cada evento (por defecto 7 días = 1 semana)."""
    eventos = []
    for r in REPERFILADOS:
        if reperfilado_afecta_eje(r, veh, eje):
            ini = pd.Timestamp(r["fecha"])
            dur = int(r.get("duracion_dias", 7))
            eventos.append((ini, ini + pd.Timedelta(days=dur)))
    return sorted(eventos, key=lambda t: t[0])


def serie_eje(df, veh, eje, param, lado):
    col = f"{param}_{'izq' if lado == 'i' else 'der'}"
    sub = df[(df["vehiculo"] == veh) & (df["eje"] == eje)].sort_values("fecha")
    sub = sub[["fecha", col]].dropna().rename(columns={col: "valor"})
    return sub


def segmento_actual(sub, veh, eje):
    """Recorta la serie al tramo posterior al último reperfilado ya
    concluido dentro de las fechas disponibles, y descarta además las
    mediciones tomadas durante la propia semana de reperfilado (son
    valores transitorios, no representativos de la tendencia)."""
    if sub.empty:
        return sub, None
    eventos = reperfilados_de_eje(veh, eje)
    for ini, fin in eventos:
        sub = sub[~((sub["fecha"] >= ini) & (sub["fecha"] <= fin))]
    if sub.empty:
        return sub, None
    concluidos = [fin for ini, fin in eventos if ini <= sub["fecha"].max()]
    if not concluidos:
        return sub, None
    corte = max(concluidos)
    return sub[sub["fecha"] > corte], corte


def forecast_wear(df, veh, eje, param, lado):
    sub = serie_eje(df, veh, eje, param, lado)
    sub_seg, corte = segmento_actual(sub, veh, eje)
    if len(sub_seg) < 2:
        if corte is not None:
            return {"sin_datos": True, "corte": corte}
        return None

    ref = sub_seg["fecha"].min()
    xs = (sub_seg["fecha"] - ref).dt.days.to_numpy(dtype=float)
    ys = sub_seg["valor"].to_numpy(dtype=float)
    if len(xs) < 2 or np.all(xs == xs[0]):
        return None

    slope, intercept = np.polyfit(xs, ys, 1)
    last_x, last_y = xs[-1], ys[-1]
    limite = LIMITES[param]

    dias_al_limite, fecha_al_limite = None, None
    if slope < -1e-6:
        dia_limite = (limite - intercept) / slope
        # Con pendientes casi planas, dia_limite puede dispararse a valores
        # absurdos (miles de años) y desbordar pandas.Timestamp. Se acota a
        # un horizonte razonable (~80 años) para evitar el overflow.
        if abs(dia_limite) < 30000:
            dias_al_limite = dia_limite - last_x
            try:
                fecha_al_limite = ref + pd.Timedelta(days=dia_limite)
            except (OverflowError, pd.errors.OutOfBoundsTimedelta):
                dias_al_limite, fecha_al_limite = None, None

    return {
        "slope": slope, "intercept": intercept, "ref": ref,
        "last_x": last_x, "last_y": last_y, "limite": limite,
        "dias_al_limite": dias_al_limite, "fecha_al_limite": fecha_al_limite,
        "corte": corte, "sub_seg": sub_seg,
    }


def motor_alertas(df, vehiculos, ejes):
    """ÚNICA fuente de verdad para el estado de alerta.

    Evalúa, para cada combinación (vehículo, eje, lado, parámetro), la
    medición REAL más reciente disponible dentro del rango de fechas
    seleccionado, contra el umbral real configurado (LIMITES) más el
    margen de alerta (MARGEN_ALERTA_PCT) — el mismo margen que ya se
    usa para colorear el heatmap. Tanto la pestaña "Alertas" como el
    resaltado de celdas en "Heatmap por tren" llaman a ESTA función:
    así ambas vistas son, por construcción, 100% coherentes entre sí.

    Estados posibles:
      - "Crítico"  → valor < límite (ya está fuera de norma)
      - "Alerta"   → límite <= valor < límite + margen (10%)
      - "Normal"   → valor >= límite + margen
      - "Sin datos"→ no hay medición para esa combinación en el rango

    La fecha reportada es siempre la fecha real de la medición que
    generó el estado (nunca la fecha de hoy ni la de carga del Excel).
    """
    filas = []
    for v in vehiculos:
        for e in ejes:
            sub_ve = df[(df["vehiculo"] == v) & (df["eje"] == e)]
            if sub_ve.empty:
                continue
            fila_ultima = sub_ve.loc[sub_ve["fecha"].idxmax()]
            for lado, suf in [("Izquierda", "izq"), ("Derecha", "der")]:
                for p in ["Sd", "Sh", "QR"]:
                    val = fila_ultima[f"{p}_{suf}"]
                    limite = LIMITES[p]
                    margen_abs = limite * MARGEN_ALERTA_PCT
                    if pd.isna(val):
                        estado = "Sin datos"
                    elif val < limite:
                        estado = "Crítico"
                    elif val < limite + margen_abs:
                        estado = "Alerta"
                    else:
                        estado = "Normal"
                    filas.append({
                        "vehiculo": v, "eje": e, "lado": lado, "parametro": p,
                        "fecha": fila_ultima["fecha"], "valor": float(val) if pd.notna(val) else None,
                        "limite": limite, "limite_alerta": limite + margen_abs,
                        "estado": estado,
                    })
    out = pd.DataFrame(filas)
    if out.empty:
        return out
    orden = {"Crítico": 0, "Alerta": 1, "Normal": 2, "Sin datos": 3}
    out["orden_estado"] = out["estado"].map(orden)
    out = out.sort_values(["orden_estado", "valor"], na_position="last").drop(columns="orden_estado")
    return out.reset_index(drop=True)


def calcular_tendencia(df, vehiculos, ejes):
    """Proyección informativa (NO es el motor de alertas): estima, a
    partir de la pendiente de desgaste reciente, cuántos días faltan
    para cruzar el límite. Es un complemento útil para priorizar
    mantenimiento, pero el estado oficial de alerta (el que se muestra
    en el heatmap y en los KPIs) siempre es el de motor_alertas()."""
    filas = []
    for v in vehiculos:
        for e in ejes:
            for lado, lado_txt in [("i", "Izquierda"), ("d", "Derecha")]:
                for p in ["Sd", "Sh", "QR"]:
                    fc = forecast_wear(df, v, e, p, lado)
                    if not fc or fc.get("sin_datos"):
                        continue
                    margen = fc["last_y"] - fc["limite"]
                    sev = None
                    if margen < 0:
                        sev = "crit"
                    elif fc["slope"] < -1e-6 and fc["dias_al_limite"] is not None and fc["dias_al_limite"] <= 180:
                        sev = "alerta"
                    elif fc["slope"] < -1e-6 and fc["dias_al_limite"] is not None and fc["dias_al_limite"] <= 400:
                        sev = "atencion"
                    if sev:
                        filas.append({
                            "vehiculo": v, "eje": e, "lado": lado_txt, "parametro": p,
                            "valor_actual": fc["last_y"], "margen": margen,
                            "dias_al_limite": fc["dias_al_limite"],
                            "fecha_al_limite": fc["fecha_al_limite"], "severidad": sev,
                        })
    if not filas:
        return pd.DataFrame()
    orden = {"crit": 0, "alerta": 1, "atencion": 2}
    out = pd.DataFrame(filas)
    out["orden_sev"] = out["severidad"].map(orden)
    out = out.sort_values(["orden_sev", "dias_al_limite"], na_position="last").drop(columns="orden_sev")
    return out


def calcular_ranking(df):
    idx = df.groupby(["vehiculo", "eje"])["fecha"].idxmax()
    ultimos = df.loc[idx]
    filas = []
    for _, r in ultimos.iterrows():
        for lado, suf in [("Izquierda", "izq"), ("Derecha", "der")]:
            for p in ["Sd", "Sh", "QR"]:
                val = r[f"{p}_{suf}"]
                if pd.isna(val):
                    continue
                filas.append({
                    "vehiculo": r["vehiculo"], "eje": r["eje"], "lado": lado,
                    "fecha": r["fecha"], "parametro": p, "valor": val,
                    "limite": LIMITES[p], "margen": val - LIMITES[p],
                })
    out = pd.DataFrame(filas).sort_values("margen")
    return out


# ============================================================
# 4. GRÁFICOS
# ============================================================
def _valor_antes_despues(serie, ini, fin):
    """Última medición antes del reperfilado y primera después de que
    termina, para poder comparar el antes/después de cada caso."""
    antes = serie[serie["fecha"] < ini]
    despues = serie[serie["fecha"] > fin]
    v_antes = float(antes.iloc[-1]["valor"]) if not antes.empty else None
    v_despues = float(despues.iloc[0]["valor"]) if not despues.empty else None
    return v_antes, v_despues


def fig_evolucion(df, veh, eje, param):
    izq = serie_eje(df, veh, eje, param, "i")
    der = serie_eje(df, veh, eje, param, "d")
    limite = LIMITES[param]

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=izq["fecha"], y=izq["valor"], mode="lines+markers",
                              name="Izquierda", line=dict(color=PAL["steel"], width=3),
                              marker=dict(size=7)))
    fig.add_trace(go.Scatter(x=der["fecha"], y=der["valor"], mode="lines+markers",
                              name="Derecha", line=dict(color=PAL["amber"], width=3),
                              marker=dict(size=7)))
    todas_fechas = pd.concat([izq["fecha"], der["fecha"]])
    if not todas_fechas.empty:
        fig.add_trace(go.Scatter(
            x=[todas_fechas.min(), todas_fechas.max()], y=[limite, limite],
            mode="lines", name=f"Límite ({limite} mm)",
            line=dict(color=PAL["red"], width=2, dash="dash"),
        ))

    for lado, serie, color in [("i", izq, PAL["steel"]), ("d", der, PAL["amber"])]:
        fc = forecast_wear(df, veh, eje, param, lado)
        if fc and not fc.get("sin_datos") and fc["slope"] < -1e-6:
            horizonte = fc["last_x"] + 200
            y_fin = fc["intercept"] + fc["slope"] * horizonte
            fig.add_trace(go.Scatter(
                x=[serie["fecha"].max(), fc["ref"] + pd.Timedelta(days=horizonte)],
                y=[fc["last_y"], y_fin], mode="lines",
                name=f"Pronóstico {'Izq' if lado == 'i' else 'Der'}",
                line=dict(color=color, width=2, dash="dot"), opacity=0.65,
            ))

    hay_eventos_visibles = False
    for ini, fin in reperfilados_de_eje(veh, eje):
        if not todas_fechas.empty and ini <= todas_fechas.max() and fin >= todas_fechas.min():
            hay_eventos_visibles = True
            fig.add_vrect(x0=ini, x1=fin, fillcolor=PAL["violet"], opacity=0.18, line_width=0,
                          annotation_text="Reperfilado", annotation_position="top left",
                          annotation_font_color=PAL["violet"], annotation_font_size=10)
            # Antes / después de ESTE reperfilado en particular, para ver
            # claramente el salto que produjo en cada lado.
            va_i, vd_i = _valor_antes_despues(izq, ini, fin)
            va_d, vd_d = _valor_antes_despues(der, ini, fin)
            partes = []
            if va_i is not None and vd_i is not None:
                partes.append(f"Izq {va_i:.2f}→{vd_i:.2f} mm (Δ{vd_i - va_i:+.2f})")
            if va_d is not None and vd_d is not None:
                partes.append(f"Der {va_d:.2f}→{vd_d:.2f} mm (Δ{vd_d - va_d:+.2f})")
            if partes:
                centro = ini + (fin - ini) / 2
                fig.add_annotation(
                    x=centro, y=1.0, yref="paper", yshift=30,
                    text="<br>".join(partes), showarrow=False, align="center",
                    font=dict(size=9.5, color=PAL["violet"]),
                    bgcolor="rgba(177,138,240,0.14)", bordercolor=PAL["violet"], borderwidth=1,
                )

    fig.update_yaxes(title=f"{param} — {LIMITES_DESC[param]} (mm)")
    fig.update_layout(title=f"Vehículo {veh} · Eje {eje}")
    fig = layout_oscuro(fig, height=460)
    # Línea de tiempo continua (todas las mediciones concatenadas
    # cronológicamente). Si hay muchas fechas o reperfilados muy juntos,
    # el rango deslizador permite acercarse a un tramo sin perder la
    # vista completa antes/después.
    fig.update_xaxes(rangeslider=dict(visible=True, thickness=0.07, bgcolor=PAL["panel2"]))
    if hay_eventos_visibles:
        fig.update_layout(margin=dict(t=95))
    return fig


def fig_heatmap_tren(df_largo, veh, param, motor_df):
    limite = LIMITES[param]
    margen_alerta = limite * MARGEN_ALERTA_PCT
    d = df_largo[(df_largo["vehiculo"] == veh) & (df_largo["variable"] == param)].copy()
    if d.empty:
        return go.Figure()

    fechas = sorted(d["fecha"].unique())
    ejes = sorted(d["eje"].unique())
    x_labels = [pd.Timestamp(f).strftime("%d/%m/%y") for f in fechas]
    y_labels = [f"Eje {e}" for e in ejes]

    def matriz(lado):
        z, txt = [], []
        for e in ejes:
            fz, ft = [], []
            for f in fechas:
                item = d[(d["eje"] == e) & (d["fecha"] == f) & (d["lado"] == lado)]
                if not item.empty:
                    v = float(item["valor"].values[0])
                    fz.append(v); ft.append(f"{v:.2f}")
                else:
                    fz.append(None); ft.append("")
            z.append(fz); txt.append(ft)
        return z, txt

    z_i, t_i = matriz("Izquierda")
    z_d, t_d = matriz("Derecha")

    valores = d["valor"].to_numpy()
    zmin, zmax = float(valores.min()) - 0.05, float(valores.max()) + 0.05

    fig = make_subplots(rows=1, cols=2, subplot_titles=["◀ Rueda Izquierda", "Rueda Derecha ▶"],
                         horizontal_spacing=0.10)
    fig.add_trace(go.Heatmap(
        z=z_i, x=x_labels, y=y_labels, zmin=zmin, zmax=zmax, colorscale=COLORSCALE_HEAT,
        text=t_i, texttemplate="%{text}", textfont={"size": 9},
        colorbar=dict(title=f"{param} (mm)", x=1.02,
                       tickvals=[limite, limite + margen_alerta, zmax],
                       ticktext=[f"{limite} mín", f"{limite+margen_alerta:.1f} alerta", f"{zmax:.1f}"]),
        xgap=2, ygap=2,
        hovertemplate="<b>%{y} — Izquierda</b><br>Fecha: %{x}<br>" + param + ": %{z:.2f} mm<extra></extra>",
    ), row=1, col=1)
    fig.add_trace(go.Heatmap(
        z=z_d, x=x_labels, y=y_labels, zmin=zmin, zmax=zmax, colorscale=COLORSCALE_HEAT,
        text=t_d, texttemplate="%{text}", textfont={"size": 9}, showscale=False,
        xgap=2, ygap=2,
        hovertemplate="<b>%{y} — Derecha</b><br>Fecha: %{x}<br>" + param + ": %{z:.2f} mm<extra></extra>",
    ), row=1, col=2)

    # ---- Coherencia con la pestaña Alertas: se usa EXACTAMENTE el mismo
    # motor_alertas() para decidir qué celdas marcar como Alerta/Crítico,
    # evaluado sobre la medición más reciente de cada eje. El color de
    # fondo es solo una ayuda visual continua; la marca ◇ es la que
    # representa el estado de alerta real (trazable 1 a 1 con la pestaña
    # "Alertas").
    motor_vp = (motor_df[(motor_df["vehiculo"] == veh) & (motor_df["parametro"] == param)]
                if motor_df is not None and not motor_df.empty else pd.DataFrame())
    n_alerta_real = int((motor_vp["estado"] == "Alerta").sum()) if not motor_vp.empty else 0
    n_crit_real = int((motor_vp["estado"] == "Crítico").sum()) if not motor_vp.empty else 0

    def _marcar_alertas(lado_nombre, col_sub):
        if motor_vp.empty:
            return
        sub = motor_vp[(motor_vp["lado"] == lado_nombre) & (motor_vp["estado"].isin(["Alerta", "Crítico"]))]
        xs, ys, colores, textos = [], [], [], []
        for _, r in sub.iterrows():
            if r["fecha"] not in fechas:
                continue  # la medición más reciente de ese eje quedó fuera del rango de fechas mostrado
            idx = fechas.index(r["fecha"])
            xs.append(x_labels[idx])
            ys.append(f"Eje {r['eje']}")
            colores.append(PAL["red"] if r["estado"] == "Crítico" else PAL["amber"])
            textos.append(
                f"<b>{r['estado'].upper()}</b><br>Eje {r['eje']} — {lado_nombre}<br>"
                f"{r['valor']:.2f} mm (límite {r['limite']:.1f} mm)<br>"
                f"Medición: {pd.Timestamp(r['fecha']).strftime('%d/%m/%Y')}"
            )
        if xs:
            fig.add_trace(go.Scatter(
                x=xs, y=ys, mode="markers",
                marker=dict(symbol="diamond-open", size=17, color=colores, line=dict(width=3)),
                hovertext=textos, hoverinfo="text", showlegend=False,
            ), row=1, col=col_sub)

    _marcar_alertas("Izquierda", 1)
    _marcar_alertas("Derecha", 2)

    # Marca en el heatmap el punto donde termina cada reperfilado de este
    # tren, para poder comparar visualmente las columnas de "antes" (a la
    # izquierda de la línea) contra las de "después" (a la derecha).
    eventos_veh = [r for r in REPERFILADOS if r["vehiculo"] == veh]
    for ev in eventos_veh:
        ini = pd.Timestamp(ev["fecha"])
        fin = ini + pd.Timedelta(days=int(ev.get("duracion_dias", 7)))
        posteriores = [f for f in fechas if f > fin]
        if not posteriores:
            continue
        label = x_labels[fechas.index(min(posteriores))]
        for col in (1, 2):
            fig.add_vline(x=label, line=dict(color=PAL["violet"], width=2, dash="dash"), row=1, col=col)
        fig.add_annotation(
            x=label, y=y_labels[0], yshift=20, row=1, col=1,
            text="🔧 antes | después", showarrow=False, font=dict(size=9, color=PAL["violet"]),
        )

    fig.update_yaxes(autorange="reversed")
    fig.update_xaxes(tickangle=-45, tickfont=dict(size=9))
    fig.update_layout(
        title=dict(
            text=(f"<b>Tren {veh}</b> — {LIMITES_DESC[param]} ({param}) · Límite mín: {limite} mm · "
                  f"Rango histórico: {valores.min():.2f}–{valores.max():.2f} mm · "
                  f"<span style='color:{PAL['amber']}'>◇ {n_alerta_real} en alerta (actual)</span> · "
                  f"<span style='color:{PAL['red']}'>◇ {n_crit_real} críticos (actual)</span>"),
            font=dict(size=13), x=0.01, xanchor="left",
        ),
    )
    return layout_oscuro(fig, height=560, legend=False, hovermode="closest")


def fig_calendario(reperfilados, fecha_min, fecha_max):
    todos = [r for r in reperfilados if r["alcance"] == "todos"]
    parcial = [r for r in reperfilados if r["alcance"] != "todos"]

    def traza(lista, nombre, color):
        return go.Bar(
            y=[f"Tren {r['vehiculo']}" for r in lista],
            # Plotly no serializa objetos pd.Timedelta a JSON: se pasa la
            # duración de la barra en milisegundos (número), que sí es
            # serializable, junto con un eje x de tipo fecha.
            x=[int(r.get("duracion_dias", 7)) * 24 * 3600 * 1000 for r in lista],
            base=[pd.Timestamp(r["fecha"]).to_pydatetime() for r in lista],
            orientation="h", name=nombre, marker=dict(color=color, line=dict(color="#fff", width=0.5)),
            text=[f"{pd.Timestamp(r['fecha']).strftime('%d/%m/%Y')} · " +
                  ("Todos los coches" if r["alcance"] == "todos" else "Solo: " + ", ".join(r["alcance"]))
                  for r in lista],
            hovertemplate="<b>%{y}</b><br>%{text}<extra></extra>",
            width=0.5,
        )

    fig = go.Figure()
    fig.add_trace(traza(todos, "Reperfilado — todos los coches", PAL["violet"]))
    fig.add_trace(traza(parcial, "Reperfilado — parcial", PAL["amber"]))
    if fecha_min is not None and fecha_max is not None:
        fig.add_vrect(x0=fecha_min, x1=fecha_max, fillcolor=PAL["steel"], opacity=0.08, line_width=0)
    fig.update_layout(barmode="overlay")
    fig.update_xaxes(type="date")
    return layout_oscuro(fig, height=280)


def fig_donut_alertas(motor_df):
    """Donut basado directamente en motor_alertas() — la MISMA fuente
    que resalta las celdas del heatmap, para garantizar coherencia."""
    if motor_df.empty:
        counts = {"Crítico": 0, "Alerta": 0, "Normal": 0}
    else:
        m = motor_df["estado"].value_counts()
        counts = {"Crítico": int(m.get("Crítico", 0)), "Alerta": int(m.get("Alerta", 0)),
                  "Normal": int(m.get("Normal", 0))}
    fig = go.Figure(go.Pie(
        labels=list(counts.keys()), values=list(counts.values()), hole=0.62,
        marker=dict(colors=[PAL["red"], PAL["amber"], PAL["green"]]),
        textinfo="value+label",
    ))
    total_alerta = counts["Crítico"] + counts["Alerta"]
    fig.add_annotation(text=f"<b>{total_alerta}</b><br>en alerta", showarrow=False,
                        font=dict(size=16, color=PAL["text"]))
    return layout_oscuro(fig, height=260, legend=False, hovermode="closest")


def fig_diametro_evolucion(df, veh, eje):
    sub = df[(df["vehiculo"] == veh) & (df["eje"] == eje)]
    izq = sub[["fecha", "kms", "Rd_izq"]].dropna().sort_values("kms")
    der = sub[["fecha", "kms", "Rd_der"]].dropna().sort_values("kms")

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=izq["kms"], y=izq["Rd_izq"], mode="lines+markers", name="Diámetro izquierda",
        line=dict(color=PAL["steel"], width=3), marker=dict(size=8),
        customdata=izq["fecha"].dt.strftime("%d/%m/%Y"),
        hovertemplate="KM: %{x}<br>Diámetro: %{y:.2f} mm<br>Fecha: %{customdata}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=der["kms"], y=der["Rd_der"], mode="lines+markers", name="Diámetro derecha",
        line=dict(color=PAL["amber"], width=3), marker=dict(size=8),
        customdata=der["fecha"].dt.strftime("%d/%m/%Y"),
        hovertemplate="KM: %{x}<br>Diámetro: %{y:.2f} mm<br>Fecha: %{customdata}<extra></extra>",
    ))
    fig.update_xaxes(title="Kilometraje (Km)")
    fig.update_yaxes(title="Diámetro de rueda — Rd (mm)")
    fig.update_layout(title=f"Evolución del diámetro · Vehículo {veh} · Eje {eje}")
    return layout_oscuro(fig, height=420)


def fig_desgaste_diametro_barras(tabla):
    """Barras del desgaste (mm) cada 10,000 km por tramo, izq vs der."""
    if tabla.empty:
        return go.Figure()
    t = tabla.copy()
    t["tramo"] = t["fecha_fin"].dt.strftime("%d/%m/%y")
    fig = go.Figure()
    for lado, color in [("Izquierda", PAL["steel"]), ("Derecha", PAL["amber"])]:
        d = t[t["lado"] == lado]
        fig.add_trace(go.Bar(
            x=d["tramo"], y=d["desgaste_cada_10000km_mm"], name=lado, marker_color=color,
            hovertemplate="Tramo hasta %{x}<br>Desgaste: %{y:.3f} mm / 10,000 km<extra></extra>",
        ))
    fig.update_layout(barmode="group", title="Desgaste de diámetro cada 10,000 km, por tramo")
    fig.update_yaxes(title="mm perdidos cada 10,000 km")
    return layout_oscuro(fig, height=380)


# ============================================================
# 5. APP STREAMLIT
# ============================================================
st.set_page_config(page_title="Dashboard Ruedas MLL2", page_icon="🚆", layout="wide")

st.markdown(f"""
<style>
    .stApp {{ background-color:{PAL['bg']}; }}
    .main-header {{
        background: linear-gradient(120deg, #12232f 0%, #1c3550 55%, #0f2a3d 100%);
        padding: 22px 30px; border-radius: 10px; margin-bottom: 18px;
        border: 1px solid {PAL['line']};
    }}
    .main-header h1 {{ color: #fff; font-size: 1.55rem; margin: 0; font-family: 'IBM Plex Sans', sans-serif; }}
    .main-header p {{ color: #a8c6df; font-size: 0.88rem; margin-top: 6px; }}
    .kpi {{
        background:{PAL['panel']}; border:1px solid {PAL['line']}; border-radius:8px;
        padding:14px 16px; text-align:left;
    }}
    .kpi .lbl {{ color:{PAL['text_dim']}; font-size:11px; text-transform:uppercase; letter-spacing:.5px; }}
    .kpi .val {{ color:{PAL['text']}; font-size:22px; font-weight:600; margin-top:2px; }}
    section[data-testid="stSidebar"] {{ background-color:{PAL['panel']}; }}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>🚆 Control de Desgaste — Ruedas MLL2</h1>
    <p>Alertas · Evolución y pronóstico · Heatmap por tren · Diámetro cada 10,000 km · Ranking · Calendario de reperfilados</p>
</div>
""", unsafe_allow_html=True)

# ---------------- Sidebar: fuente de datos ----------------
st.sidebar.header("📂 Fuente de datos")
archivo_subido = st.sidebar.file_uploader(
    "Subir Excel (.xlsx) o CSV con la misma estructura de columnas",
    type=["xlsx", "csv"],
)

try:
    if archivo_subido is not None:
        es_csv = archivo_subido.name.lower().endswith(".csv")
        df_ancho, n_omitidas, n_duplicadas = cargar_desde_bytes(archivo_subido.name, archivo_subido.getvalue(), es_csv)
        st.sidebar.success(f"Usando archivo subido: **{archivo_subido.name}**")
    else:
        ruta_defecto = localizar_excel_por_defecto()
        if ruta_defecto is None:
            st.sidebar.warning("No hay archivo en la carpeta **datos**. Sube uno arriba para empezar.")
            st.info("⬅️ Sube un archivo Excel o CSV en la barra lateral para generar el dashboard.")
            st.stop()
        es_csv = ruta_defecto.suffix.lower() == ".csv"
        df_ancho, n_omitidas, n_duplicadas = cargar_desde_ruta(str(ruta_defecto), es_csv)
        st.sidebar.info(f"Usando archivo por defecto:\n`{ruta_defecto.name}`")
except Exception as e:
    st.error(
        f"❌ No se pudo procesar el archivo: {e}\n\n"
        "Verifica que mantenga las columnas: acción, vehículo, fecha, kms, eje, "
        "Sd/Qr/Sh/Rd izquierda, Sd/Qr/Sh/Rd derecha."
    )
    st.stop()

if n_omitidas:
    st.sidebar.caption(f"⚠️ {n_omitidas} fila(s) del archivo se omitieron por datos incompletos o mal formados.")
if n_duplicadas:
    st.sidebar.caption(
        f"⚠️ {n_duplicadas} fila(s) con el mismo vehículo+eje+fecha — se conservó la última de cada grupo."
    )

# ---------------- Sidebar: rango de fechas (calendario) ----------------
st.sidebar.header("🗓️ Rango de fechas")
fecha_min_data = df_ancho["fecha"].min().date()
fecha_max_data = df_ancho["fecha"].max().date()
rango = st.sidebar.date_input(
    "Selecciona el rango a analizar",
    value=(fecha_min_data, fecha_max_data),
    min_value=fecha_min_data, max_value=fecha_max_data,
)
if isinstance(rango, tuple) and len(rango) == 2:
    f_ini, f_fin = rango
else:
    f_ini, f_fin = fecha_min_data, fecha_max_data

df = df_ancho[
    (df_ancho["fecha"].dt.date >= f_ini) & (df_ancho["fecha"].dt.date <= f_fin)
].copy()

if df.empty:
    st.warning("No hay mediciones dentro del rango de fechas seleccionado.")
    st.stop()

df_largo = a_formato_largo(df)
vehiculos = sorted(df["vehiculo"].unique().tolist())
ejes_todos = sorted(df["eje"].unique().tolist())

# ---------------- Sidebar: selección tren / eje / parámetro ----------------
st.sidebar.header("🔧 Selección")
veh_sel = st.sidebar.selectbox("Vehículo (tren)", vehiculos)
ejes_veh = sorted(df[df["vehiculo"] == veh_sel]["eje"].unique().tolist())
eje_sel = st.sidebar.selectbox("Eje", ejes_veh)
param_sel = st.sidebar.radio("Parámetro", ["Sd", "Sh", "QR"], horizontal=True,
                              format_func=lambda p: f"{p} — {LIMITES_DESC[p]}")

# ---------------- Sidebar: editor manual de fechas de reperfilado ----------------
st.sidebar.header("🛠️ Reperfilados")
with st.sidebar.expander("✏️ Editar fechas de reperfilado", expanded=False):
    st.caption(
        "Agrega, edita o elimina filas. **Alcance**: escribe `Todos`, o los "
        "coches separados por coma (ej. `M1, M2`) si el reperfilado fue parcial. "
        "**Duración**: días aproximados que dura (por defecto 7 = 1 semana); "
        "las mediciones tomadas dentro de esa ventana se ignoran en las "
        "tendencias, porque son valores transitorios."
    )
    tabla_reperf_base = pd.DataFrame([
        {
            "Vehículo": r["vehiculo"],
            "Fecha": pd.Timestamp(r["fecha"]).date(),
            "Duración (días)": int(r.get("duracion_dias", 7)),
            "Alcance": "Todos" if r["alcance"] == "todos" else ", ".join(r["alcance"]),
        }
        for r in REPERFILADOS_DEFAULT
    ])
    tabla_reperf_editada = st.data_editor(
        tabla_reperf_base, num_rows="dynamic", use_container_width=True, hide_index=True,
        key="editor_reperfilados",
        column_config={
            "Vehículo": st.column_config.NumberColumn(step=1, format="%d"),
            "Duración (días)": st.column_config.NumberColumn(step=1, format="%d", min_value=1),
        },
    )

_nuevos_reperf = []
for _, r in tabla_reperf_editada.iterrows():
    if pd.isna(r.get("Vehículo")) or pd.isna(r.get("Fecha")):
        continue
    try:
        veh_r = int(r["Vehículo"])
    except (TypeError, ValueError):
        continue
    alcance_txt = str(r.get("Alcance", "Todos")).strip()
    if alcance_txt.lower() in ("", "todos", "all", "nan"):
        alcance_r = "todos"
    else:
        alcance_r = [c.strip().upper() for c in alcance_txt.split(",") if c.strip()]
    dur_r = r.get("Duración (días)", 7)
    dur_r = int(dur_r) if pd.notna(dur_r) else 7
    _nuevos_reperf.append({
        "vehiculo": veh_r, "fecha": pd.Timestamp(r["Fecha"]),
        "duracion_dias": dur_r, "alcance": alcance_r,
    })
# Se reemplaza la lista "viva" que usan todas las funciones de análisis
# (reperfilado_afecta_eje, forecast_wear, fig_evolucion, fig_calendario, etc.)
REPERFILADOS = _nuevos_reperf if _nuevos_reperf else list(REPERFILADOS_DEFAULT)

# ---------------- Sidebar: mapeo eje → coche (para reperfilados parciales) ----------------
with st.sidebar.expander("🚃 Ejes por coche (M1 / M2 / ...)", expanded=False):
    st.caption(
        "El Excel no indica a qué coche pertenece cada eje, así que hace falta "
        "definirlo aquí SOLO para poder aplicar correctamente los reperfilados "
        "**parciales** (los generales — 'Todos' — no lo necesitan). "
        "Columna **Ejes**: por ejemplo `1-12` o `1,2,3,4`. "
        "Mientras un tren con reperfilado parcial no tenga aquí sus ejes "
        "mapeados, ese evento parcial NO se aplicará a ningún eje de ese tren "
        "(para no filtrar por error datos de un coche que no fue reperfilado)."
    )
    tabla_map_base = pd.DataFrame([{"Vehículo": None, "Coche": "", "Ejes": ""}])
    tabla_map_editada = st.data_editor(
        tabla_map_base, num_rows="dynamic", use_container_width=True, hide_index=True,
        key="editor_coches",
        column_config={"Vehículo": st.column_config.NumberColumn(step=1, format="%d")},
    )


def _parsear_rango_ejes(txt):
    out = set()
    for parte in str(txt).split(","):
        parte = parte.strip()
        if not parte:
            continue
        if "-" in parte:
            a, b = parte.split("-", 1)
            try:
                a, b = int(a.strip()), int(b.strip())
                out.update(range(min(a, b), max(a, b) + 1))
            except ValueError:
                continue
        else:
            try:
                out.add(int(parte))
            except ValueError:
                continue
    return out


EJE_COCHE_MAP = {}
for _, r in tabla_map_editada.iterrows():
    if pd.isna(r.get("Vehículo")) or not str(r.get("Coche", "")).strip() or not str(r.get("Ejes", "")).strip():
        continue
    try:
        veh_m = int(r["Vehículo"])
    except (TypeError, ValueError):
        continue
    coche_m = str(r["Coche"]).strip().upper()
    for e in _parsear_rango_ejes(r["Ejes"]):
        EJE_COCHE_MAP[f"{veh_m}|{e}"] = coche_m

_faltan_mapeo = hay_reperfilados_parciales_sin_mapear(vehiculos, ejes_todos)
if _faltan_mapeo:
    st.sidebar.warning(
        "⚠️ Hay reperfilados **parciales** para el/los tren(es) "
        f"{', '.join('Tren ' + str(v) for v in _faltan_mapeo)} pero no tienen "
        "ejes mapeados a coche todavía → ese evento no se está aplicando a "
        "ningún eje. Complétalo en '🚃 Ejes por coche' arriba."
    )

# ---------------- Motor de alertas (única fuente de verdad) ----------------
motor_df = motor_alertas(df, vehiculos, ejes_todos)
n_alerta_kpi = int((motor_df["estado"] == "Alerta").sum()) if not motor_df.empty else 0
n_crit = int((motor_df["estado"] == "Crítico").sum()) if not motor_df.empty else 0

# ---------------- KPIs ----------------
k1, k2, k3, k4, k5 = st.columns(5)
for col, lbl, val in [
    (k1, "Mediciones", f"{len(df):,}"),
    (k2, "Vehículos", f"{len(vehiculos)}"),
    (k3, "Rango", f"{f_ini.strftime('%d/%m/%y')} – {f_fin.strftime('%d/%m/%y')}"),
    (k4, "En alerta", f"{n_alerta_kpi}"),
    (k5, "Críticas", f"{n_crit}"),
]:
    col.markdown(f'<div class="kpi"><div class="lbl">{lbl}</div><div class="val">{val}</div></div>',
                 unsafe_allow_html=True)

st.write("")

# ---------------- Tabs ----------------
tab_alertas, tab_evol, tab_heat, tab_diam, tab_rank, tab_cal = st.tabs(
    ["🚨 Alertas", "📈 Evolución por eje", "🔥 Heatmap por tren",
     "📏 Diámetro cada 10,000 km", "🏆 Ranking", "🗓️ Calendario de reperfilados"]
)

with tab_alertas:
    st.caption(
        "Estado calculado sobre la **medición real más reciente** de cada eje, contra el "
        "límite normativo (± margen de alerta del 10%). Esta es la misma fuente que resalta "
        "las celdas ◇ en el Heatmap — ambas vistas son, por construcción, coherentes entre sí."
    )
    c1, c2 = st.columns([1, 2])
    with c1:
        st.plotly_chart(fig_donut_alertas(motor_df), use_container_width=True)
    with c2:
        criticos_alertas = motor_df[motor_df["estado"].isin(["Crítico", "Alerta"])].copy() if not motor_df.empty else pd.DataFrame()
        if criticos_alertas.empty:
            st.success("Sin alertas — ninguna rueda está actualmente por debajo del margen de seguridad.")
        else:
            criticos_alertas["fecha"] = criticos_alertas["fecha"].dt.strftime("%d/%m/%Y")
            criticos_alertas["estado"] = criticos_alertas["estado"].map(
                {"Crítico": "🔴 CRÍTICO", "Alerta": "🟠 ALERTA"})
            mostrar = criticos_alertas.rename(columns={
                "vehiculo": "Vehículo", "eje": "Eje", "lado": "Lado", "parametro": "Parámetro",
                "fecha": "Fecha medición", "valor": "Valor (mm)", "limite": "Límite (mm)",
                "limite_alerta": "Límite + margen (mm)", "estado": "Estado",
            })[["Vehículo", "Eje", "Lado", "Parámetro", "Fecha medición", "Valor (mm)",
                "Límite (mm)", "Límite + margen (mm)", "Estado"]]
            st.dataframe(mostrar, use_container_width=True, hide_index=True)

    with st.expander("📉 Proyección de tendencia (informativa — no determina el estado de alerta)"):
        st.caption(
            "Estimación de cuándo cada rueda cruzaría el límite si mantiene su ritmo de desgaste "
            "actual, según regresión sobre las mediciones desde el último reperfilado concluido. "
            "Es un apoyo para priorizar mantenimiento, pero el estado oficial de alerta (arriba, "
            "y en el Heatmap) siempre es el de la medición real más reciente, no una proyección."
        )
        tendencia_df = calcular_tendencia(df, vehiculos, ejes_todos)
        if tendencia_df.empty:
            st.info("No hay tendencias de desgaste proyectables en el rango seleccionado.")
        else:
            mostrar_t = tendencia_df.head(20).copy()
            mostrar_t["severidad"] = mostrar_t["severidad"].map(
                {"crit": "🔴 Ya bajo el límite", "alerta": "🟠 < 180 días", "atencion": "🟡 < 400 días"})
            mostrar_t["fecha_al_limite"] = mostrar_t["fecha_al_limite"].dt.strftime("%d/%m/%Y")
            mostrar_t["dias_al_limite"] = mostrar_t["dias_al_limite"].round(0)
            mostrar_t = mostrar_t.rename(columns={
                "vehiculo": "Vehículo", "eje": "Eje", "lado": "Lado", "parametro": "Parámetro",
                "valor_actual": "Valor actual (mm)", "margen": "Margen (mm)",
                "dias_al_limite": "Días al límite (proyectado)", "fecha_al_limite": "Fecha estimada",
                "severidad": "Horizonte",
            })
            st.dataframe(mostrar_t, use_container_width=True, hide_index=True)

with tab_evol:
    st.plotly_chart(fig_evolucion(df, veh_sel, eje_sel, param_sel), use_container_width=True)

    fc_i = forecast_wear(df, veh_sel, eje_sel, param_sel, "i")
    fc_d = forecast_wear(df, veh_sel, eje_sel, param_sel, "d")
    lineas = []
    for lado_txt, fc in [("Izquierda", fc_i), ("Derecha", fc_d)]:
        if not fc:
            lineas.append(f"**{lado_txt}:** datos insuficientes para proyectar tendencia.")
        elif fc.get("sin_datos"):
            lineas.append(f"**{lado_txt}:** solo hay 1 medición desde el fin del reperfilado del "
                           f"{fc['corte'].strftime('%d/%m/%Y')}; se necesita otra medición para proyectar.")
        elif fc["slope"] >= -1e-6:
            lineas.append(f"**{lado_txt}:** tendencia estable / sin desgaste progresivo detectado.")
        elif fc["dias_al_limite"] is not None and fc["dias_al_limite"] < 0:
            lineas.append(f"**{lado_txt}:** valor ya por debajo del límite de {fc['limite']} mm.")
        elif fc["dias_al_limite"] is None:
            lineas.append(f"**{lado_txt}:** tendencia decreciente muy leve; no se puede estimar una fecha confiable.")
        else:
            lineas.append(
                f"**{lado_txt}:** desgaste de {abs(fc['slope']):.5f} mm/día → alcanzaría el límite "
                f"({fc['limite']} mm) en **~{round(fc['dias_al_limite'])} días** "
                f"(≈ {fc['fecha_al_limite'].strftime('%d/%m/%Y')})."
            )
    st.markdown(f"""<div class="kpi">{'<br>'.join(lineas)}</div>""", unsafe_allow_html=True)
    st.caption("La banda morada sombreada marca la semana aproximada de cada reperfilado; "
               "esas mediciones se excluyen del cálculo de tendencia.")

with tab_heat:
    veh_heat = st.selectbox("Tren para el heatmap", vehiculos, key="veh_heat",
                             index=vehiculos.index(veh_sel))
    param_heat = st.radio("Parámetro", ["Sd", "Sh", "QR"], horizontal=True, key="param_heat",
                           format_func=lambda p: f"{p} — {LIMITES_DESC[p]}")
    st.plotly_chart(fig_heatmap_tren(df_largo, veh_heat, param_heat, motor_df), use_container_width=True)
    st.caption("Línea vertical morada = primer valor medido después de terminado el reperfilado de "
               "ese tren (columnas a la izquierda = antes, a la derecha = después). "
               "Marca ◇ ámbar/roja = estado Alerta/Crítico según la medición más reciente de ese eje "
               "(la misma fuente que la pestaña Alertas).")
    with st.expander("🔍 Ver todos los trenes a la vez"):
        for v in vehiculos:
            st.plotly_chart(fig_heatmap_tren(df_largo, v, param_heat, motor_df), use_container_width=True,
                             key=f"heat_all_{v}")

with tab_diam:
    st.markdown(f"##### Vehículo {veh_sel} · Eje {eje_sel}")
    st.plotly_chart(fig_diametro_evolucion(df, veh_sel, eje_sel), use_container_width=True)

    tabla_diam = calcular_desgaste_diametro(df, veh_sel, eje_sel)
    if tabla_diam.empty:
        st.info("No hay al menos 2 mediciones de diámetro (Rd) para este eje en el rango seleccionado.")
    else:
        st.plotly_chart(fig_desgaste_diametro_barras(tabla_diam), use_container_width=True)
        mostrar_d = tabla_diam.copy()
        mostrar_d["fecha_inicio"] = mostrar_d["fecha_inicio"].dt.strftime("%d/%m/%Y")
        mostrar_d["fecha_fin"] = mostrar_d["fecha_fin"].dt.strftime("%d/%m/%Y")
        mostrar_d = mostrar_d.rename(columns={
            "lado": "Lado", "fecha_inicio": "Fecha inicio", "fecha_fin": "Fecha fin",
            "km_inicio": "Km inicio", "km_fin": "Km fin", "delta_km": "ΔKm",
            "diametro_inicio_mm": "Diámetro inicio (mm)", "diametro_fin_mm": "Diámetro fin (mm)",
            "delta_diametro_mm": "ΔDiámetro (mm)",
            "desgaste_cada_10000km_mm": "Desgaste cada 10,000 km (mm)",
        })
        st.dataframe(mostrar_d, use_container_width=True, hide_index=True)
        st.caption("ΔDiámetro negativo = la rueda perdió diámetro (desgaste normal). "
                   "Un ΔDiámetro positivo suele indicar que hubo un reperfilado no registrado en el calendario.")

    diff_lr = calcular_diferencia_izq_der(df, veh_sel, eje_sel)
    if not diff_lr.empty:
        st.markdown("##### Diferencia de diámetro Izquierda − Derecha (mismo eje, mismo día)")
        d2 = diff_lr.copy()
        d2["fecha"] = d2["fecha"].dt.strftime("%d/%m/%Y")
        d2 = d2.rename(columns={
            "fecha": "Fecha", "kms": "Km", "Rd_izq": "Diámetro izq. (mm)",
            "Rd_der": "Diámetro der. (mm)", "diferencia_izq_der_mm": "Diferencia Izq−Der (mm)",
        })
        st.dataframe(d2, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("##### 🏆 Ranking — mayor desgaste de diámetro (todos los trenes y ejes, en el rango)")
    global_d = calcular_desgaste_diametro_global(df, vehiculos, ejes_todos)
    if global_d.empty:
        st.info("No hay suficientes datos de diámetro (Rd) para generar este ranking.")
    else:
        rank_d = global_d.sort_values("desgaste_cada_10000km_mm", ascending=False).head(15).copy()
        rank_d["fecha_fin"] = rank_d["fecha_fin"].dt.strftime("%d/%m/%Y")
        rank_d.insert(0, "#", range(1, len(rank_d) + 1))
        rank_d = rank_d.rename(columns={
            "vehiculo": "Vehículo", "eje": "Eje", "lado": "Lado", "fecha_fin": "Fecha",
            "km_fin": "Km", "desgaste_cada_10000km_mm": "Desgaste cada 10,000 km (mm)",
        })[["#", "Vehículo", "Eje", "Lado", "Fecha", "Km", "Desgaste cada 10,000 km (mm)"]]
        st.dataframe(
            rank_d, use_container_width=True, hide_index=True,
            column_config={
                "Desgaste cada 10,000 km (mm)": st.column_config.ProgressColumn(
                    "Desgaste cada 10,000 km (mm)", min_value=0.0,
                    max_value=float(max(rank_d["Desgaste cada 10,000 km (mm)"].max(), 0.1)),
                    format="%.3f",
                )
            },
        )

with tab_rank:
    ranking_df = calcular_ranking(df).head(15).copy()
    if ranking_df.empty:
        st.info("No hay datos suficientes para el ranking en el rango seleccionado.")
    else:
        ranking_df["fecha"] = ranking_df["fecha"].dt.strftime("%d/%m/%Y")
        ranking_df.insert(0, "#", range(1, len(ranking_df) + 1))
        ranking_df = ranking_df.rename(columns={
            "vehiculo": "Vehículo", "eje": "Eje", "lado": "Lado", "fecha": "Fecha",
            "parametro": "Parámetro", "valor": "Valor (mm)", "limite": "Límite (mm)", "margen": "Margen (mm)",
        })
        st.dataframe(
            ranking_df, use_container_width=True, hide_index=True,
            column_config={
                "Margen (mm)": st.column_config.ProgressColumn(
                    "Margen (mm)", min_value=float(min(ranking_df["Margen (mm)"].min(), 0)),
                    max_value=3.5, format="%.2f",
                )
            },
        )

with tab_cal:
    fecha_min_ts = pd.Timestamp(f_ini)
    fecha_max_ts = pd.Timestamp(f_fin)
    st.plotly_chart(fig_calendario(REPERFILADOS, fecha_min_ts, fecha_max_ts), use_container_width=True)
    st.caption("Cada barra representa la semana aproximada del reperfilado. "
               "Morado = todos los coches del tren · Ámbar = reperfilado parcial.")

    tabla_cal = []
    for r in sorted(REPERFILADOS, key=lambda x: x["fecha"]):
        f = pd.Timestamp(r["fecha"])
        dentro = fecha_min_ts <= f <= fecha_max_ts
        tabla_cal.append({
            "Fecha": f.strftime("%d/%m/%Y"), "Vehículo": f"Tren {r['vehiculo']}",
            "Duración (días)": int(r.get("duracion_dias", 7)),
            "Alcance": "Todos los coches" if r["alcance"] == "todos" else "Solo: " + ", ".join(r["alcance"]),
            "En rango seleccionado": "✅ Sí" if dentro else "— No",
        })
    st.dataframe(pd.DataFrame(tabla_cal), use_container_width=True, hide_index=True)

st.markdown("---")
st.caption(
    f"Límites de referencia — Sd ≥ 25.0 mm · Sh ≥ 27.5 mm · QR ≥ 6.5 mm · "
    f"{len(df):,} mediciones cargadas · {len(vehiculos)} vehículos en el rango {f_ini} – {f_fin}"
)
